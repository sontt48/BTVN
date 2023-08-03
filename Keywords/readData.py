import openpyxl
import os
current_working_directory = os.getcwd()
data = current_working_directory + "\Keywords\Book1.xlsx"
open_data = openpyxl.load_workbook(data)
sheet1_data = open_data["Sheet1"]


def  get_data(row,column):
    cell = sheet1_data.cell(row,column)
    return cell.value

def get_data_row_number():
    return sheet1_data.max_row

def  get_data_column_number():
    return sheet1_data.max_column

print(get_data(1,2))

